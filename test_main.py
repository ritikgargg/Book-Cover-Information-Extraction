from main import main
from authorsExtraction import extractAuthors
from publishersExtraction import extractPublishers
from isbnsExtraction import extractISBNs
from fileValidator import inputFileValidator

import openpyxl


def read_excel_cell(worksheet, row, col):
    '''To read the value of a cell in an excel worksheet'''

    cell_obj = worksheet.cell(row=row, column=col)
    return cell_obj.value


def test_main_single_file():
    '''Unit test for the case when a single input file is given as input'''

    argv = ['main.py', '0', 'images\\image2.png']
    main(argv)

    # Reading the resultant excel file, and validating the results
    workbook = openpyxl.load_workbook('output.xlsx')
    worksheet = workbook['Sheet1']
    assert worksheet.max_row == 2
    assert worksheet.max_column == 4
    assert read_excel_cell(worksheet, 2, 1) == "ERAGON"


def test_main_directory():
    '''Unit test for the case when a directory of files is given as input'''

    argv = ['main.py', '1', 'images']
    main(argv)

    # Reading the resultant excel file, and validating the results
    workbook = openpyxl.load_workbook('output.xlsx')
    worksheet = workbook['Sheet1']
    assert worksheet.max_row == 5
    assert worksheet.max_column == 4

    assert read_excel_cell(
        worksheet, 2, 1) == "Excellence: Can We Be Equal and Excellent Too?"
    assert read_excel_cell(worksheet, 3, 1) == "ERAGON"
    assert read_excel_cell(worksheet, 4, 1) == "COMPUTER NETWORKING"
    assert read_excel_cell(worksheet, 5, 1) == "Clean Code"


def test_extractAuthors():
    imgText = "Computer Networking is written by James F. Kurose and Keith W. Ross"
    authors = extractAuthors(imgText)
    assert "James F. Kurose" in authors
    assert "Keith W. Ross" in authors


def test_extractPublishers():
    imgText = "Pearson Publications is a reknowned book publishing company"
    publishers = extractPublishers(imgText)
    assert "Pearson Publications" in publishers


def test_extractISBNs():
    imgText = "Computer Networking has ISBN-10: 0-13-285620-4 and ISBN-13: 978-0-13-285620-1"
    isbns = extractISBNs(imgText)
    assert "ISBN-10: 0-13-285620-4" in isbns
    assert "ISBN-13: 978-0-13-285620-1" in isbns


def test_fileValidator():
    path = "test.pdf"
    assert inputFileValidator(path) == False
